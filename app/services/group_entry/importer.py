"""
Importador de links de grupos de CSV/Excel.

Sprint 25 - E12 - S12.1

Suporta:
- CSV com colunas: name, url, invite_code, state, category
- Excel (.xlsx) com mesma estrutura
- Importação de diretório completo
"""

import csv
import re
import logging
from pathlib import Path
from typing import Optional, Tuple, List
from datetime import datetime, UTC

from pydantic import BaseModel

from app.services.supabase import supabase

logger = logging.getLogger(__name__)


class GroupLink(BaseModel):
    """Link de grupo para importação."""

    invite_code: str
    nome: Optional[str] = None
    invite_url: Optional[str] = None
    categoria: Optional[str] = None
    estado: Optional[str] = None
    regiao: Optional[str] = None


def extrair_invite_code(url: str) -> Optional[str]:
    """
    Extrai invite code de uma URL de grupo.

    Formatos suportados:
    - https://chat.whatsapp.com/ABC123...
    - ABC123... (código direto)

    Args:
        url: URL ou código do grupo

    Returns:
        Código de convite extraído ou None
    """
    if not url:
        return None

    url = str(url).strip()

    # Se for URL completa
    if "chat.whatsapp.com/" in url:
        parts = url.split("chat.whatsapp.com/")
        if len(parts) > 1:
            return parts[1].split("?")[0].strip()

    # Se for código direto (alfanumérico, ~22 chars)
    if len(url) >= 20 and url.replace("-", "").replace("_", "").isalnum():
        return url

    return None


def validar_formato_invite_code(invite_code: str) -> bool:
    """
    Valida formato do invite code ANTES de inserir no banco.
    Evita poluir DB com links obviamente inválidos.

    Args:
        invite_code: Código a validar

    Returns:
        True se formato válido
    """
    if not invite_code:
        return False

    # Formato esperado: alfanumérico com - e _, 20-24 caracteres
    if not re.match(r"^[A-Za-z0-9_-]{18,26}$", invite_code):
        return False

    return True


async def verificar_duplicado(invite_code: str) -> bool:
    """
    Verifica se link já existe no banco.

    Args:
        invite_code: Código do convite

    Returns:
        True se já existe
    """
    result = (
        supabase.table("group_links")
        .select("id")
        .eq("invite_code", invite_code)
        .execute()
    )

    return len(result.data) > 0 if result.data else False


async def _validar_link_antes_importar(invite_code: str) -> Tuple[bool, str]:
    """
    Validação completa ANTES de inserir no banco.

    Args:
        invite_code: Código do convite

    Returns:
        (valido, motivo)
    """
    # 1. Formato válido?
    if not validar_formato_invite_code(invite_code):
        return False, "formato_invalido"

    # 2. Já existe no banco?
    if await verificar_duplicado(invite_code):
        return False, "duplicado"

    return True, "ok"


async def importar_csv(
    arquivo: Path,
    categoria: Optional[str] = None,
    estado: Optional[str] = None,
) -> dict:
    """
    Importa links de um arquivo CSV.

    Args:
        arquivo: Caminho do arquivo CSV
        categoria: Categoria padrão (sobrescreve se não houver coluna)
        estado: Estado padrão

    Returns:
        {
            "total_linhas": N,
            "importados": N,
            "duplicados": N,
            "invalidos": N,
            "erros": [...]
        }
    """
    if not arquivo.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")

    resultado = {
        "total_linhas": 0,
        "importados": 0,
        "duplicados": 0,
        "invalidos": 0,
        "erros": [],
    }

    links_para_inserir = []

    with open(arquivo, encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for i, row in enumerate(reader, 1):
            resultado["total_linhas"] += 1

            # Extrair invite code
            invite_code = extrair_invite_code(
                row.get("invite_code") or row.get("url") or row.get("link", "")
            )

            if not invite_code:
                resultado["invalidos"] += 1
                resultado["erros"].append(f"Linha {i}: invite_code inválido")
                continue

            # Validar formato
            if not validar_formato_invite_code(invite_code):
                resultado["invalidos"] += 1
                resultado["erros"].append(f"Linha {i}: formato inválido")
                continue

            link = GroupLink(
                invite_code=invite_code,
                nome=row.get("name") or row.get("nome"),
                invite_url=row.get("url") or row.get("link"),
                categoria=row.get("category") or row.get("categoria") or categoria,
                estado=row.get("state") or row.get("estado") or estado,
                regiao=row.get("regiao") or row.get("region"),
            )

            links_para_inserir.append(
                {
                    "invite_code": link.invite_code,
                    "invite_url": link.invite_url,
                    "nome": link.nome,
                    "categoria": link.categoria,
                    "estado": link.estado,
                    "regiao": link.regiao,
                    "fonte": arquivo.name,
                    "status": "pendente",
                }
            )

    # Inserir em batch (ignorando duplicados)
    if links_para_inserir:
        try:
            result = (
                supabase.table("group_links")
                .upsert(
                    links_para_inserir,
                    on_conflict="invite_code",
                    ignore_duplicates=True,
                )
                .execute()
            )

            resultado["importados"] = len(result.data) if result.data else 0
            resultado["duplicados"] = len(links_para_inserir) - resultado["importados"]

        except Exception as e:
            logger.error(f"Erro ao inserir links: {e}")
            resultado["erros"].append(f"Erro no banco: {str(e)}")

    logger.info(
        f"[GroupImporter] {arquivo.name}: "
        f"{resultado['importados']} importados, "
        f"{resultado['duplicados']} duplicados, "
        f"{resultado['invalidos']} inválidos"
    )

    return resultado


async def importar_excel(
    arquivo: Path,
    sheet_name: Optional[str] = None,
    categoria: Optional[str] = None,
    estado: Optional[str] = None,
) -> dict:
    """
    Importa links de um arquivo Excel (.xlsx).

    Espera colunas: name, url/invite_code, state, category

    Args:
        arquivo: Caminho do arquivo Excel
        sheet_name: Nome da planilha (opcional, usa ativa)
        categoria: Categoria padrão
        estado: Estado padrão

    Returns:
        Resultado da importação
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl é necessário para importar Excel. Instale com: uv add openpyxl")

    if not arquivo.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")

    wb = load_workbook(arquivo, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    resultado = {
        "total_linhas": 0,
        "importados": 0,
        "duplicados": 0,
        "invalidos": 0,
        "erros": [],
    }

    links_para_inserir = []
    headers = None

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Primeira linha = headers
        if i == 1:
            headers = [str(h).lower().strip() if h else "" for h in row]
            continue

        resultado["total_linhas"] += 1

        # Montar dict da linha
        row_dict = dict(zip(headers, row))

        # Extrair invite code
        invite_code = extrair_invite_code(
            str(
                row_dict.get("invite_code")
                or row_dict.get("url")
                or row_dict.get("link", "")
            )
        )

        if not invite_code:
            resultado["invalidos"] += 1
            continue

        # Validar formato
        if not validar_formato_invite_code(invite_code):
            resultado["invalidos"] += 1
            continue

        links_para_inserir.append(
            {
                "invite_code": invite_code,
                "invite_url": str(row_dict.get("url") or row_dict.get("link") or ""),
                "nome": str(row_dict.get("name") or row_dict.get("nome") or ""),
                "categoria": str(
                    row_dict.get("category")
                    or row_dict.get("categoria")
                    or categoria
                    or ""
                ),
                "estado": str(
                    row_dict.get("state") or row_dict.get("estado") or estado or ""
                ),
                "regiao": str(row_dict.get("regiao") or row_dict.get("region") or ""),
                "fonte": arquivo.name,
                "status": "pendente",
            }
        )

    wb.close()

    # Inserir em batch
    if links_para_inserir:
        try:
            result = (
                supabase.table("group_links")
                .upsert(
                    links_para_inserir,
                    on_conflict="invite_code",
                    ignore_duplicates=True,
                )
                .execute()
            )

            resultado["importados"] = len(result.data) if result.data else 0
            resultado["duplicados"] = len(links_para_inserir) - resultado["importados"]

        except Exception as e:
            logger.error(f"Erro ao inserir links: {e}")
            resultado["erros"].append(f"Erro no banco: {str(e)}")

    logger.info(
        f"[GroupImporter] {arquivo.name}: "
        f"{resultado['importados']} importados, "
        f"{resultado['duplicados']} duplicados, "
        f"{resultado['invalidos']} inválidos"
    )

    return resultado


async def importar_diretorio(diretorio: Path) -> dict:
    """
    Importa todos os CSVs e Excel de um diretório.

    Args:
        diretorio: Caminho do diretório

    Returns:
        Resultado consolidado da importação
    """
    resultado_total = {
        "arquivos_processados": 0,
        "total_importados": 0,
        "total_duplicados": 0,
        "total_invalidos": 0,
        "detalhes": [],
    }

    for arquivo in diretorio.glob("*.csv"):
        resultado = await importar_csv(arquivo)
        resultado_total["arquivos_processados"] += 1
        resultado_total["total_importados"] += resultado["importados"]
        resultado_total["total_duplicados"] += resultado["duplicados"]
        resultado_total["total_invalidos"] += resultado["invalidos"]
        resultado_total["detalhes"].append({"arquivo": arquivo.name, **resultado})

    for arquivo in diretorio.glob("*.xlsx"):
        resultado = await importar_excel(arquivo)
        resultado_total["arquivos_processados"] += 1
        resultado_total["total_importados"] += resultado["importados"]
        resultado_total["total_duplicados"] += resultado["duplicados"]
        resultado_total["total_invalidos"] += resultado["invalidos"]
        resultado_total["detalhes"].append({"arquivo": arquivo.name, **resultado})

    logger.info(
        f"[GroupImporter] Diretório {diretorio}: "
        f"{resultado_total['arquivos_processados']} arquivos, "
        f"{resultado_total['total_importados']} links importados"
    )

    return resultado_total


async def listar_links(
    status: Optional[str] = None,
    categoria: Optional[str] = None,
    limite: int = 100,
    offset: int = 0,
) -> List[dict]:
    """
    Lista links de grupos com filtros.

    Args:
        status: Filtrar por status
        categoria: Filtrar por categoria
        limite: Limite de resultados
        offset: Offset para paginação

    Returns:
        Lista de links
    """
    query = supabase.table("group_links").select("*")

    if status:
        query = query.eq("status", status)

    if categoria:
        query = query.eq("categoria", categoria)

    result = query.order("created_at", desc=True).range(offset, offset + limite - 1).execute()

    return result.data or []


async def contar_links_por_status() -> dict:
    """
    Conta links por status.

    Returns:
        Dict com contagem por status
    """
    result = supabase.rpc("count_group_links_by_status").execute()

    if result.data:
        return {row["status"]: row["count"] for row in result.data}

    # Fallback: query manual
    result = supabase.table("group_links").select("status").execute()

    if not result.data:
        return {}

    contagem = {}
    for row in result.data:
        status = row["status"]
        contagem[status] = contagem.get(status, 0) + 1

    return contagem
